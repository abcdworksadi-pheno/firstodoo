# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
import json


class LettreMotivationExcelSource(models.Model):
    """Source Excel pour les lettres de motivation (niveau 3)"""
    _name = 'lettre.motivation.excel.source'
    _description = 'Source Excel pour Lettre de Motivation'
    _order = 'name'

    name = fields.Char(
        string='Nom de la Source',
        required=True,
        help='Nom de la source Excel'
    )

    fichier_excel = fields.Binary(
        string='Fichier Excel',
        required=True,
        help='Fichier Excel (.xlsx)'
    )

    nom_fichier = fields.Char(
        string='Nom du Fichier',
        help='Nom du fichier Excel'
    )

    feuille = fields.Char(
        string='Nom de la Feuille',
        default='Sheet1',
        help='Nom de la feuille Excel à utiliser'
    )

    range_cellules = fields.Char(
        string='Plage de Cellules',
        help='Plage de cellules à lire (ex: A1:D10). Laisser vide pour toute la feuille.'
    )

    mapping_colonnes = fields.Text(
        string='Mapping des Colonnes',
        help='Mapping JSON: {"colonne_excel": "colonne_tableau"}'
    )

    mapping_colonnes_dict = fields.Char(
        string='Mapping (Dict)',
        compute='_compute_mapping_dict',
        store=False
    )

    tableau_id = fields.Many2one(
        'lettre.motivation.tableau',
        string='Tableau Associé',
        help='Tableau qui sera rempli avec les données Excel'
    )

    active = fields.Boolean(
        string='Actif',
        default=True
    )

    @api.depends('mapping_colonnes')
    def _compute_mapping_dict(self):
        """Convertit le mapping JSON en dictionnaire"""
        for record in self:
            if record.mapping_colonnes:
                try:
                    record.mapping_colonnes_dict = json.dumps(
                        json.loads(record.mapping_colonnes),
                        ensure_ascii=False,
                        indent=2
                    )
                except:
                    record.mapping_colonnes_dict = record.mapping_colonnes
            else:
                record.mapping_colonnes_dict = '{}'

    def get_mapping_dict(self):
        """Retourne le mapping sous forme de dictionnaire Python"""
        if not self.mapping_colonnes:
            return {}
        try:
            return json.loads(self.mapping_colonnes)
        except:
            return {}
    
    def _get_column_letter(self, col_idx):
        """Convertit un index de colonne (1, 2, 3...) en lettre (A, B, C...)"""
        result = ""
        while col_idx > 0:
            col_idx -= 1
            result = chr(65 + (col_idx % 26)) + result
            col_idx //= 26
        return result

    def action_importer_excel(self):
        """Importe les données Excel dans le tableau associé"""
        self.ensure_one()
        
        if not self.fichier_excel:
            raise UserError(_('Aucun fichier Excel chargé!'))
        
        if not self.tableau_id:
            raise UserError(_('Aucun tableau associé!'))
        
        try:
            import base64
            import io
            from openpyxl import load_workbook
            
            # Décoder le fichier
            excel_data = base64.b64decode(self.fichier_excel)
            workbook = load_workbook(io.BytesIO(excel_data))
            
            # Sélectionner la feuille
            if self.feuille not in workbook.sheetnames:
                raise UserError(_('La feuille "%s" n\'existe pas dans le fichier!') % self.feuille)
            
            sheet = workbook[self.feuille]
            
            # Déterminer la plage de cellules
            if self.range_cellules:
                # TODO: Parser la plage (ex: A1:D10) - pour l'instant on utilise toute la feuille
                max_row = sheet.max_row
                max_col = sheet.max_column
            else:
                # Utiliser toute la feuille
                max_row = sheet.max_row
                max_col = sheet.max_column
            
            # Lire les données
            mapping = self.get_mapping_dict()
            colonnes_tableau = {col.name: col for col in self.tableau_id.colonnes_ids}
            
            # Supprimer les anciennes lignes
            self.tableau_id.lignes_ids.unlink()
            
            # Lire les lignes (en supposant que la première ligne contient les en-têtes)
            lignes_data = []
            
            # Lire la première ligne pour obtenir les en-têtes
            headers = []
            for col_idx in range(1, max_col + 1):
                header_cell = sheet.cell(row=1, column=col_idx).value
                if header_cell:
                    # Convertir l'index de colonne en lettre (A, B, C...)
                    col_letter = self._get_column_letter(col_idx)
                    headers.append((col_letter, str(header_cell)))
            
            # Lire les données
            for row_idx in range(2, max_row + 1):
                row_values = []
                for col_idx in range(1, max_col + 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    row_values.append(cell_value)
                
                if not any(v is not None for v in row_values):  # Ignorer les lignes vides
                    continue
                
                valeurs = {}
                for col_idx, (col_letter, header) in enumerate(headers):
                    if col_idx < len(row_values):
                        cell_value = row_values[col_idx]
                        # Mapper la colonne Excel vers la colonne du tableau
                        col_name = mapping.get(col_letter, header.lower().replace(' ', '_'))
                        if col_name in colonnes_tableau:
                            valeurs[col_name] = str(cell_value) if cell_value else ''
                
                if valeurs:
                    lignes_data.append({
                        'tableau_id': self.tableau_id.id,
                        'valeurs': json.dumps(valeurs, ensure_ascii=False),
                        'sequence': row_idx - 1,
                    })
            
            # Créer les lignes
            if lignes_data:
                self.env['lettre.motivation.tableau.ligne'].create(lignes_data)
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Import réussi'),
                    'message': _('%d lignes importées avec succès!') % len(lignes_data),
                    'type': 'success',
                    'sticky': False,
                }
            }
            
        except ImportError:
            raise UserError(_('La bibliothèque openpyxl n\'est pas installée! Installez-la avec: pip install openpyxl'))
        except Exception as e:
            raise UserError(_('Erreur lors de l\'import Excel: %s') % str(e))

